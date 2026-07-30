from __future__ import annotations

import json
from dataclasses import dataclass
from io import BytesIO
from typing import Any
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session, joinedload

from .database import get_db
from .models import Fragrance, FragranceNote

router = APIRouter(prefix="/api/ai-research-import", tags=["ai-research-import"])

REQUIRED_SHEETS = {
    "Düfte", "Noten", "Performance", "Duft-DNA", "Bilder_Quellen",
    "Preisquellen", "Quellen", "Anleitung", "Metadaten",
}

PERSONAL_FIELDS = {
    "personal_longevity_hours", "personal_projection", "personal_sillage",
    "personal_performance_score", "personal_fragrance_dna",
}

SHEET_FIELDS: dict[str, tuple[str, ...]] = {
    "Düfte": (
        "dgd_id", "year", "gender", "concentration", "perfumer", "price_eur",
        "description", "accords", "sweetness", "freshness",
    ),
    "Performance": (
        "longevity", "projection", "longevity_min_hours", "longevity_max_hours",
        "longevity_score", "sillage", "performance_score", "projection_first_hour",
        "projection_after_three_hours", "drydown_strength", "performance_source_count",
        "performance_confidence", "performance_disagreement", "performance_status",
        "performance_version", "performance_production_period",
    ),
    "Duft-DNA": (
        "fragrance_dna_json", "fragrance_dna_source", "fragrance_dna_status",
        "fragrance_dna_source_count", "fragrance_dna_confidence",
        "fragrance_dna_disagreement",
    ),
    "Bilder_Quellen": (
        "image_url", "image_source_name", "image_source_url", "image_usage_note",
        "image_status",
    ),
}

JSON_FIELDS = {"fragrance_dna_json"}
INTEGER_FIELDS = {"year", "performance_source_count", "fragrance_dna_source_count"}
FLOAT_FIELDS = {
    "price_eur", "sweetness", "freshness", "longevity", "projection",
    "longevity_min_hours", "longevity_max_hours", "longevity_score", "sillage",
    "performance_score", "projection_first_hour", "projection_after_three_hours",
    "drydown_strength", "performance_confidence", "performance_disagreement",
    "fragrance_dna_confidence", "fragrance_dna_disagreement",
}


@dataclass
class ParsedWorkbook:
    export_id: str
    schema_version: str
    rows: dict[str, list[dict[str, Any]]]


def _text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _normalize(field: str, value: Any) -> Any:
    if _is_blank(value):
        return None
    if field in JSON_FIELDS:
        if isinstance(value, (dict, list)):
            return value
        try:
            return json.loads(str(value))
        except json.JSONDecodeError as exc:
            raise ValueError(f"Ungültiges JSON in {field}: {exc.msg}") from exc
    if field in INTEGER_FIELDS:
        return int(float(value))
    if field in FLOAT_FIELDS:
        return float(value)
    return str(value).strip()


def _serializable(value: Any) -> Any:
    if isinstance(value, UUID):
        return str(value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _same(left: Any, right: Any) -> bool:
    if left is None and right is None:
        return True
    if isinstance(left, (dict, list)) or isinstance(right, (dict, list)):
        return left == right
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return abs(float(left) - float(right)) < 1e-9
    return str(left or "").strip() == str(right or "").strip()


def _sheet_rows(ws) -> list[dict[str, Any]]:
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return []
    headers = [_text(item) for item in values[0]]
    if not headers or any(not item for item in headers):
        raise HTTPException(400, f"Tabellenblatt {ws.title} enthält leere Spaltennamen.")
    if len(headers) != len(set(headers)):
        raise HTTPException(400, f"Tabellenblatt {ws.title} enthält doppelte Spaltennamen.")
    rows: list[dict[str, Any]] = []
    for raw in values[1:]:
        row = {headers[index]: raw[index] if index < len(raw) else None for index in range(len(headers))}
        if any(not _is_blank(value) for value in row.values()):
            rows.append(row)
    return rows


def _parse_workbook(content: bytes) -> ParsedWorkbook:
    try:
        workbook = load_workbook(BytesIO(content), data_only=False, read_only=True)
    except Exception as exc:
        raise HTTPException(400, "Die Datei ist keine lesbare XLSX-Arbeitsmappe.") from exc

    missing = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
    if missing:
        raise HTTPException(400, f"Pflichtblätter fehlen: {', '.join(missing)}")

    metadata_rows = _sheet_rows(workbook["Metadaten"])
    metadata = {_text(row.get("key")): _text(row.get("value")) for row in metadata_rows}
    export_id = metadata.get("export_id", "")
    schema_version = metadata.get("schema_version", "")
    if not export_id:
        raise HTTPException(400, "In Metadaten fehlt export_id.")
    if not schema_version.startswith("16.7"):
        raise HTTPException(400, f"Nicht unterstützte schema_version: {schema_version or 'leer'}")

    rows = {name: _sheet_rows(workbook[name]) for name in REQUIRED_SHEETS if name not in {"Anleitung", "Metadaten"}}
    for sheet_name, sheet_rows in rows.items():
        for row_number, row in enumerate(sheet_rows, start=2):
            row_export_id = _text(row.get("export_id"))
            if row_export_id and row_export_id != export_id:
                raise HTTPException(400, f"{sheet_name}, Zeile {row_number}: export_id wurde verändert.")
            forbidden = PERSONAL_FIELDS.intersection(row)
            if forbidden:
                raise HTTPException(400, f"{sheet_name}: persönliche Felder sind nicht erlaubt: {', '.join(sorted(forbidden))}")
    return ParsedWorkbook(export_id=export_id, schema_version=schema_version, rows=rows)


def _fragrance_map(db: Session, ids: set[UUID]) -> dict[UUID, Fragrance]:
    if not ids:
        return {}
    items = (
        db.query(Fragrance)
        .options(joinedload(Fragrance.brand), joinedload(Fragrance.note_links).joinedload(FragranceNote.note))
        .filter(Fragrance.id.in_(ids))
        .all()
    )
    return {item.id: item for item in items}


def _collect_ids(parsed: ParsedWorkbook) -> set[UUID]:
    result: set[UUID] = set()
    for sheet, rows in parsed.rows.items():
        for row_number, row in enumerate(rows, start=2):
            raw = _text(row.get("fragrance_id"))
            if not raw:
                if sheet in {"Quellen", "Preisquellen"}:
                    continue
                raise HTTPException(400, f"{sheet}, Zeile {row_number}: fragrance_id fehlt.")
            try:
                result.add(UUID(raw))
            except ValueError as exc:
                raise HTTPException(400, f"{sheet}, Zeile {row_number}: ungültige fragrance_id.") from exc
    return result


def _field_preview(parsed: ParsedWorkbook, fragrances: dict[UUID, Fragrance]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    changes: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for sheet, fields in SHEET_FIELDS.items():
        for row_number, row in enumerate(parsed.rows.get(sheet, []), start=2):
            fragrance_id = UUID(_text(row.get("fragrance_id")))
            fragrance = fragrances.get(fragrance_id)
            if not fragrance:
                errors.append({"sheet": sheet, "row": row_number, "message": "fragrance_id existiert nicht."})
                continue
            for field in fields:
                if field not in row or _is_blank(row.get(field)):
                    continue
                try:
                    new_value = _normalize(field, row.get(field))
                except (ValueError, TypeError) as exc:
                    errors.append({"sheet": sheet, "row": row_number, "field": field, "message": str(exc)})
                    continue
                model_field = "fragrance_dna" if field == "fragrance_dna_json" else field
                old_value = getattr(fragrance, model_field, None)
                if _same(old_value, new_value):
                    continue
                changes.append({
                    "key": f"{sheet}:{fragrance_id}:{field}",
                    "sheet": sheet,
                    "row": row_number,
                    "fragrance_id": str(fragrance_id),
                    "fragrance_name": fragrance.name,
                    "brand_name": fragrance.brand.name if fragrance.brand else "",
                    "field": field,
                    "old_value": _serializable(old_value),
                    "new_value": _serializable(new_value),
                    "kind": "new" if _is_blank(old_value) else "conflict",
                })
    return changes, errors


def _note_preview(parsed: ParsedWorkbook, fragrances: dict[UUID, Fragrance]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    changes: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for row_number, row in enumerate(parsed.rows.get("Noten", []), start=2):
        fragrance_id = UUID(_text(row.get("fragrance_id")))
        fragrance = fragrances.get(fragrance_id)
        if not fragrance:
            errors.append({"sheet": "Noten", "row": row_number, "message": "fragrance_id existiert nicht."})
            continue
        note_name = _text(row.get("note_name"))
        pyramid = _text(row.get("pyramid")).upper()
        if not note_name or not pyramid:
            continue
        try:
            position = int(float(row.get("position") or 0))
        except (TypeError, ValueError):
            errors.append({"sheet": "Noten", "row": row_number, "field": "position", "message": "Position muss eine Zahl sein."})
            continue
        exists = any(
            link.note and link.note.name.strip().casefold() == note_name.casefold()
            and link.pyramid.strip().upper() == pyramid
            and int(link.position or 0) == position
            for link in fragrance.note_links
        )
        if not exists:
            changes.append({
                "key": f"Noten:{fragrance_id}:{pyramid}:{position}:{note_name.casefold()}",
                "sheet": "Noten", "row": row_number, "fragrance_id": str(fragrance_id),
                "fragrance_name": fragrance.name,
                "brand_name": fragrance.brand.name if fragrance.brand else "",
                "field": f"{pyramid}-Note #{position}", "old_value": None,
                "new_value": note_name, "kind": "new", "preview_only": True,
            })
    return changes, errors


def _price_preview(parsed: ParsedWorkbook, fragrances: dict[UUID, Fragrance]) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    for row_number, row in enumerate(parsed.rows.get("Preisquellen", []), start=2):
        if _is_blank(row.get("product_url")):
            continue
        fragrance_id = UUID(_text(row.get("fragrance_id")))
        fragrance = fragrances.get(fragrance_id)
        if not fragrance:
            continue
        changes.append({
            "key": f"Preisquellen:{fragrance_id}:{row_number}",
            "sheet": "Preisquellen", "row": row_number, "fragrance_id": str(fragrance_id),
            "fragrance_name": fragrance.name,
            "brand_name": fragrance.brand.name if fragrance.brand else "",
            "field": "Produktlink", "old_value": None,
            "new_value": {
                "merchant_name": _text(row.get("merchant_name")),
                "product_url": _text(row.get("product_url")),
                "product_variant": _text(row.get("product_variant")),
                "size_ml": row.get("size_ml"),
                "concentration": _text(row.get("concentration")),
                "product_kind": _text(row.get("product_kind")),
                "current_price": row.get("current_price"),
                "currency": _text(row.get("currency")) or "EUR",
            },
            "kind": "new", "preview_only": True,
        })
    return changes


@router.post("/preview")
async def preview_ai_research_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Bitte eine XLSX-Datei auswählen.")
    content = await file.read()
    if not content or len(content) > 25 * 1024 * 1024:
        raise HTTPException(400, "Die Datei ist leer oder größer als 25 MB.")

    parsed = _parse_workbook(content)
    ids = _collect_ids(parsed)
    fragrances = _fragrance_map(db, ids)
    unknown_ids = sorted(str(item) for item in ids - set(fragrances))
    if unknown_ids:
        raise HTTPException(400, f"Unbekannte fragrance_id: {', '.join(unknown_ids[:5])}")

    field_changes, field_errors = _field_preview(parsed, fragrances)
    note_changes, note_errors = _note_preview(parsed, fragrances)
    price_changes = _price_preview(parsed, fragrances)
    changes = field_changes + note_changes + price_changes
    errors = field_errors + note_errors

    summary = {
        "fragrances": len(ids),
        "changes": len(changes),
        "new_values": sum(1 for item in changes if item["kind"] == "new"),
        "conflicts": sum(1 for item in changes if item["kind"] == "conflict"),
        "preview_only": sum(1 for item in changes if item.get("preview_only")),
        "errors": len(errors),
    }
    return {
        "filename": filename,
        "export_id": parsed.export_id,
        "schema_version": parsed.schema_version,
        "summary": summary,
        "changes": changes,
        "errors": errors,
        "database_changed": False,
    }
