from __future__ import annotations

import json
from datetime import datetime, timezone
from io import BytesIO
from uuid import UUID, uuid4

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.orm import Session, joinedload

from .database import get_db
from .models import Fragrance

router = APIRouter(prefix="/api/ai-research-export", tags=["ai-research-export"])

PERSONAL_FIELDS = {
    "personal_longevity_hours", "personal_projection", "personal_sillage",
    "personal_performance_score", "personal_fragrance_dna",
}

FRAGRANCE_HEADERS = [
    "export_id", "fragrance_id", "dgd_id", "brand_id", "brand_name", "name",
    "year", "gender", "concentration", "perfumer", "price_eur", "description",
    "accords", "sweetness", "freshness", "created_at",
]
PERFORMANCE_HEADERS = [
    "export_id", "fragrance_id", "brand_name", "name", "longevity", "projection",
    "longevity_min_hours", "longevity_max_hours", "longevity_score", "sillage",
    "performance_score", "projection_first_hour", "projection_after_three_hours",
    "drydown_strength", "performance_source_count", "performance_confidence",
    "performance_disagreement", "performance_status", "performance_researched_at",
    "performance_version", "performance_production_period", "proposal_sources_json",
    "proposal_source_url", "proposal_rationale", "proposal_confidence",
]
DNA_HEADERS = [
    "export_id", "fragrance_id", "brand_name", "name", "fragrance_dna_json",
    "fragrance_dna_source", "fragrance_dna_status", "fragrance_dna_source_count",
    "fragrance_dna_confidence", "fragrance_dna_disagreement",
    "fragrance_dna_researched_at", "proposal_source_label", "proposal_source_url",
    "proposal_rationale", "proposal_confidence", "proposal_values_json",
]
NOTE_HEADERS = [
    "export_id", "fragrance_id", "brand_name", "name", "pyramid", "position",
    "note_id", "note_name", "note_category", "note_description",
]
IMAGE_HEADERS = [
    "export_id", "fragrance_id", "brand_name", "name", "image_url",
    "image_source_name", "image_source_url", "image_usage_note", "image_status",
]
PRICE_SOURCE_HEADERS = [
    "export_id", "fragrance_id", "offer_source_id", "brand_name", "name",
    "merchant_name", "product_url", "product_variant", "size_ml", "concentration",
    "product_kind", "current_price", "shipping_cost", "total_price", "currency",
    "availability", "ean_gtin", "merchant_sku", "price_per_100ml", "market_country",
    "last_checked_at", "scanner_active", "scan_interval", "extraction_hint",
    "trust_status", "variant_warning",
]
SOURCE_HEADERS = [
    "export_id", "fragrance_id", "source_id", "source_name", "source_url",
    "source_type", "source_date", "usage_status", "trust_status", "field_scope", "note",
]


def _json(value) -> str:
    if value in (None, ""):
        return ""
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _iso(value) -> str:
    return value.isoformat() if value else ""


def _style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3A342F")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for column in range(1, ws.max_column + 1):
        width = 14
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=1, max_row=min(ws.max_row, 100)):
            for item in cell:
                width = min(max(width, len(str(item.value or "")) + 2), 48)
        ws.column_dimensions[get_column_letter(column)].width = width


def _append(ws, headers: list[str], values: dict) -> None:
    ws.append([values.get(header, "") for header in headers])


def _optional_rows(db: Session, statement: str, params: dict | None = None) -> list[dict]:
    try:
        return [dict(row) for row in db.execute(text(statement), params or {}).mappings().all()]
    except Exception:
        db.rollback()
        return []


def _has_missing(fragrance: Fragrance) -> bool:
    required = (
        fragrance.year, fragrance.concentration, fragrance.perfumer, fragrance.description,
        fragrance.image_url, fragrance.image_source_url, fragrance.accords,
        fragrance.fragrance_dna, fragrance.longevity_min_hours, fragrance.longevity_max_hours,
        fragrance.sillage,
    )
    return any(value in (None, "", [], {}) for value in required) or not fragrance.note_links


@router.get("/xlsx")
def export_ai_research_xlsx(
    scope: str = Query(default="missing", pattern="^(missing|all)$"),
    brand_id: UUID | None = None,
    fragrance_ids: str | None = None,
    db: Session = Depends(get_db),
):
    export_id = str(uuid4())
    selected_ids = {item.strip() for item in (fragrance_ids or "").split(",") if item.strip()}
    query = db.query(Fragrance).options(
        joinedload(Fragrance.brand), joinedload(Fragrance.note_links).joinedload("note")
    ).order_by(Fragrance.name)
    if brand_id:
        query = query.filter(Fragrance.brand_id == brand_id)
    fragrances = query.all()
    if selected_ids:
        fragrances = [item for item in fragrances if str(item.id) in selected_ids]
    if scope == "missing":
        fragrances = [item for item in fragrances if _has_missing(item)]

    fragrance_id_list = [str(item.id) for item in fragrances]
    dna_proposals = _optional_rows(db, """
        SELECT fragrance_id::text, values, source_label, source_url, rationale, confidence
        FROM fragrance_dna_proposals WHERE status = 'OPEN'
    """)
    performance_proposals = _optional_rows(db, """
        SELECT fragrance_id::text, values, sources, source_url, rationale, confidence
        FROM performance_research_proposals WHERE status = 'OPEN'
    """)
    sources = _optional_rows(db, """
        SELECT id, object_id, name, file_or_url, source_type, source_date,
               usage_status, trust_status, object_type, note
        FROM master_sources
        WHERE object_type ILIKE '%%fragrance%%' OR object_type IS NULL
    """)

    dna_by_id = {row["fragrance_id"]: row for row in dna_proposals}
    perf_by_id = {row["fragrance_id"]: row for row in performance_proposals}

    wb = Workbook()
    wb.remove(wb.active)
    sheets = {
        "Düfte": FRAGRANCE_HEADERS,
        "Noten": NOTE_HEADERS,
        "Performance": PERFORMANCE_HEADERS,
        "Duft-DNA": DNA_HEADERS,
        "Bilder_Quellen": IMAGE_HEADERS,
        "Preisquellen": PRICE_SOURCE_HEADERS,
        "Quellen": SOURCE_HEADERS,
    }
    for title, headers in sheets.items():
        ws = wb.create_sheet(title)
        ws.append(headers)

    for fragrance in fragrances:
        common = {
            "export_id": export_id, "fragrance_id": str(fragrance.id),
            "brand_name": fragrance.brand.name if fragrance.brand else "", "name": fragrance.name,
        }
        _append(wb["Düfte"], FRAGRANCE_HEADERS, {
            **common, "dgd_id": fragrance.dgd_id or "", "brand_id": str(fragrance.brand_id),
            "year": fragrance.year or "", "gender": fragrance.gender or "",
            "concentration": fragrance.concentration or "", "perfumer": fragrance.perfumer or "",
            "price_eur": fragrance.price_eur if fragrance.price_eur is not None else "",
            "description": fragrance.description or "", "accords": fragrance.accords or "",
            "sweetness": fragrance.sweetness if fragrance.sweetness is not None else "",
            "freshness": fragrance.freshness if fragrance.freshness is not None else "",
            "created_at": _iso(fragrance.created_at),
        })
        _append(wb["Bilder_Quellen"], IMAGE_HEADERS, {
            **common, "image_url": fragrance.image_url or "",
            "image_source_name": fragrance.image_source_name or "",
            "image_source_url": fragrance.image_source_url or "",
            "image_usage_note": fragrance.image_usage_note or "",
            "image_status": fragrance.image_status or "",
        })
        perf = perf_by_id.get(str(fragrance.id), {})
        _append(wb["Performance"], PERFORMANCE_HEADERS, {
            **common, **{key: getattr(fragrance, key, "") if getattr(fragrance, key, None) is not None else "" for key in PERFORMANCE_HEADERS},
            "performance_researched_at": _iso(fragrance.performance_researched_at),
            "proposal_sources_json": _json(perf.get("sources")),
            "proposal_source_url": perf.get("source_url") or "",
            "proposal_rationale": perf.get("rationale") or "",
            "proposal_confidence": perf.get("confidence") if perf.get("confidence") is not None else "",
        })
        dna = dna_by_id.get(str(fragrance.id), {})
        _append(wb["Duft-DNA"], DNA_HEADERS, {
            **common, "fragrance_dna_json": _json(fragrance.fragrance_dna),
            "fragrance_dna_source": fragrance.fragrance_dna_source or "",
            "fragrance_dna_status": fragrance.fragrance_dna_status or "",
            "fragrance_dna_source_count": fragrance.fragrance_dna_source_count or "",
            "fragrance_dna_confidence": fragrance.fragrance_dna_confidence if fragrance.fragrance_dna_confidence is not None else "",
            "fragrance_dna_disagreement": fragrance.fragrance_dna_disagreement if fragrance.fragrance_dna_disagreement is not None else "",
            "fragrance_dna_researched_at": _iso(fragrance.fragrance_dna_researched_at),
            "proposal_source_label": dna.get("source_label") or "",
            "proposal_source_url": dna.get("source_url") or "",
            "proposal_rationale": dna.get("rationale") or "",
            "proposal_confidence": dna.get("confidence") if dna.get("confidence") is not None else "",
            "proposal_values_json": _json(dna.get("values")),
        })
        for link in sorted(fragrance.note_links, key=lambda item: (item.pyramid, item.position)):
            _append(wb["Noten"], NOTE_HEADERS, {
                **common, "pyramid": link.pyramid, "position": link.position,
                "note_id": str(link.note_id), "note_name": link.note.name if link.note else "",
                "note_category": link.note.category if link.note else "",
                "note_description": link.note.description if link.note else "",
            })

    allowed = set(fragrance_id_list)
    for source in sources:
        object_id = str(source.get("object_id") or "")
        if object_id and object_id not in allowed:
            continue
        _append(wb["Quellen"], SOURCE_HEADERS, {
            "export_id": export_id, "fragrance_id": object_id,
            "source_id": source.get("id") or "", "source_name": source.get("name") or "",
            "source_url": source.get("file_or_url") or "", "source_type": source.get("source_type") or "",
            "source_date": _iso(source.get("source_date")), "usage_status": source.get("usage_status") or "",
            "trust_status": source.get("trust_status") or "", "field_scope": source.get("object_type") or "",
            "note": source.get("note") or "",
        })

    instruction = wb.create_sheet("Anleitung")
    instruction.append(["KI-Rechercheanleitung"])
    rules = [
        "Nur belegbare Werte ergänzen; nichts erfinden.",
        "Vorhandene Werte nicht löschen oder überschreiben, sofern keine belastbare Korrektur samt Quelle vorliegt.",
        "Leere Zellen bedeuten keine Löschung.",
        "fragrance_id, export_id und vorhandene offer_source_id niemals verändern.",
        "Pro recherchierter Feldgruppe mindestens einen direkten Quellenlink angeben.",
        "Bei Preisquellen ausschließlich direkte Produktseiten verwenden, keine Such- oder Kategorieseiten.",
        "Größe, Konzentration, Tester/Set, Währung und Markt eindeutig erfassen.",
        "Performance-Reformulierung und Produktionszeitraum getrennt behandeln.",
        "JSON-Spalten als valides JSON belassen.",
        "Tabellenblätter und technische Spaltennamen nicht umbenennen.",
    ]
    for rule in rules:
        instruction.append([rule])
    metadata = wb.create_sheet("Metadaten")
    metadata.append(["key", "value"])
    metadata.append(["export_id", export_id])
    metadata.append(["created_at", datetime.now(timezone.utc).isoformat()])
    metadata.append(["scope", scope])
    metadata.append(["fragrance_count", len(fragrances)])
    metadata.append(["schema_version", "16.7.1"])
    metadata.append(["personal_fields_excluded", ",".join(sorted(PERSONAL_FIELDS))])

    for ws in wb.worksheets:
        _style_sheet(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"DGD_KI_Recherche_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
