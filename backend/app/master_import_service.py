from __future__ import annotations

import re
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, time
from io import BytesIO
from typing import Any, Iterable
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import Brand, Fragrance, MasterImportRun, MasterPerfumer, MasterSource, TwinMatch


class MasterImportValidationError(ValueError):
    pass


@dataclass
class Counter:
    created: int = 0
    updated: int = 0
    unchanged: int = 0


@dataclass
class MasterImportReport:
    filename: str
    detected_version: str | None = None
    dry_run: bool = True
    brands: Counter = field(default_factory=Counter)
    fragrances: Counter = field(default_factory=Counter)
    twins: Counter = field(default_factory=Counter)
    perfumers: Counter = field(default_factory=Counter)
    sources: Counter = field(default_factory=Counter)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        raw = asdict(self)
        raw.update({
            "total_rows": sum(
                raw[name][key]
                for name in ("brands", "fragrances", "twins", "perfumers", "sources")
                for key in ("created", "updated", "unchanged")
            ),
            "valid_rows": sum(
                raw[name][key]
                for name in ("brands", "fragrances", "twins", "perfumers", "sources")
                for key in ("created", "updated", "unchanged")
            ),
            "duplicate_count": sum(raw[name]["unchanged"] for name in ("brands", "fragrances", "twins", "perfumers", "sources")),
            "error_count": len(self.errors),
            "warning_count": len(self.warnings),
            "new_brand_count": self.brands.created,
            "new_fragrance_count": self.fragrances.created,
            "new_twin_count": self.twins.created,
            "new_perfumer_count": self.perfumers.created,
            "new_source_count": self.sources.created,
            "created": self.brands.created + self.fragrances.created + self.twins.created + self.perfumers.created + self.sources.created,
            "updated": self.brands.updated + self.fragrances.updated + self.twins.updated + self.perfumers.updated + self.sources.updated,
            "skipped": self.brands.unchanged + self.fragrances.unchanged + self.twins.unchanged + self.perfumers.unchanged + self.sources.unchanged,
            "failed": len(self.errors),
        })
        return raw


SHEET_ALIASES = {
    "brands": ("Marken", "Brands"),
    "fragrances": ("Düfte", "Duefte", "Fragrances"),
    "twins": ("Duftzwillinge", "Dupes"),
    "sources": ("Quellen", "Sources"),
    "perfumers": ("Parfümeure", "Parfumeure", "Perfumers"),
}


def _clean(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if isinstance(value, datetime):
        return value
    return value


def _normal(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _int(value: Any) -> int | None:
    value = _clean(value)
    if value is None:
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _float(value: Any) -> float | None:
    value = _clean(value)
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _bool(value: Any) -> bool | None:
    value = _clean(value)
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    folded = str(value).casefold()
    if folded in {"ja", "yes", "true", "1", "x"}:
        return True
    if folded in {"nein", "no", "false", "0"}:
        return False
    return None


def _date(value: Any, epoch=None) -> date | None:
    value = _clean(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and epoch is not None:
        try:
            return from_excel(value, epoch).date()
        except Exception:
            pass
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value), fmt).date()
        except ValueError:
            pass
    return None


def _datetime(value: Any, epoch=None) -> datetime | None:
    parsed = _date(value, epoch)
    return datetime.combine(parsed, time.min) if parsed else None


def _json_value(value: Any, epoch=None) -> Any:
    value = _clean(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and epoch is not None:
        return value
    return value


def _json_row(data: dict[str, Any], epoch=None) -> dict[str, Any]:
    return {str(key): _json_value(value, epoch) for key, value in data.items() if key and value is not None}


def _find_sheet(workbook, aliases: Iterable[str]):
    folded = {name.casefold(): name for name in workbook.sheetnames}
    for alias in aliases:
        actual = folded.get(alias.casefold())
        if actual:
            return workbook[actual]
    return None


def _rows(sheet) -> Iterable[tuple[int, dict[str, Any]]]:
    headers = [_clean(cell.value) for cell in sheet[1]]
    last_header = max((index for index, header in enumerate(headers) if header is not None), default=-1)
    headers = headers[: last_header + 1]
    for row_number, cells in enumerate(
        sheet.iter_rows(min_row=2, max_col=len(headers), values_only=True),
        start=2,
    ):
        values = [_clean(value) for value in cells]
        if not any(value is not None for value in values):
            continue
        yield row_number, dict(zip(headers, values))


def _detect_version(workbook) -> str | None:
    for sheet_name in ("README", "Dashboard"):
        if sheet_name not in workbook.sheetnames:
            continue
        for row in workbook[sheet_name].iter_rows(min_row=1, max_row=4, values_only=True):
            for value in row:
                if isinstance(value, str):
                    match = re.search(r"v(\d+(?:\.\d+)*)", value, re.I)
                    if match:
                        return match.group(1)
    return None


def _gender(value: Any) -> str:
    folded = _normal(value)
    if "damen" in folded and "herren" in folded:
        return "Unisex"
    if "damen" in folded:
        return "Damen"
    if "herren" in folded:
        return "Herren"
    return "Unisex"


def _changed(existing: Any, values: dict[str, Any]) -> bool:
    return any(getattr(existing, key) != value for key, value in values.items())


def _fill_only(existing: Any, key: str, value: Any, values: dict[str, Any]) -> None:
    if value is not None and getattr(existing, key) in (None, ""):
        values[key] = value


def _prepare_workbook(file_bytes: bytes, filename: str):
    if not filename.lower().endswith(".xlsx"):
        raise MasterImportValidationError("Für die Master-Datenbank wird eine XLSX-Datei benötigt.")
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise MasterImportValidationError(f"Die Excel-Datei konnte nicht gelesen werden: {exc}") from exc

    sheets = {key: _find_sheet(workbook, aliases) for key, aliases in SHEET_ALIASES.items()}
    missing = [name for name in ("brands", "fragrances", "twins") if sheets[name] is None]
    if missing:
        raise MasterImportValidationError(f"Pflichtblätter fehlen: {', '.join(missing)}")
    return workbook, sheets


def _validate(workbook, sheets):
    errors: list[str] = []
    brands: dict[str, dict[str, Any]] = {}
    fragrances: dict[str, dict[str, Any]] = {}
    twins: dict[str, dict[str, Any]] = {}
    perfumers: dict[str, dict[str, Any]] = {}
    sources: dict[str, dict[str, Any]] = {}

    for row_number, data in _rows(sheets["brands"]):
        key, name = data.get("Marken-ID"), data.get("Marke")
        if not key or not name:
            errors.append(f"Marken!{row_number}: Marken-ID oder Marke fehlt.")
            continue
        key = str(key)
        if key in brands:
            errors.append(f"Marken!{row_number}: doppelte Marken-ID {key}.")
        brands[key] = data

    for row_number, data in _rows(sheets["fragrances"]):
        key, brand_key, name = data.get("DGD-ID"), data.get("Marken-ID"), data.get("Duftname")
        if not key or not brand_key or not name:
            errors.append(f"Düfte!{row_number}: DGD-ID, Marken-ID oder Duftname fehlt.")
            continue
        key, brand_key = str(key), str(brand_key)
        if key in fragrances:
            errors.append(f"Düfte!{row_number}: doppelte DGD-ID {key}.")
        if brand_key not in brands:
            errors.append(f"Düfte!{row_number}: unbekannte Marken-ID {brand_key}.")
        fragrances[key] = data

    seen_pairs: set[tuple[str, str]] = set()
    for row_number, data in _rows(sheets["twins"]):
        key = data.get("Zuordnungs-ID")
        original = data.get("Original-DGD-ID")
        alternative = data.get("Alternative-DGD-ID")
        if not key or not original or not alternative:
            errors.append(f"Duftzwillinge!{row_number}: Zuordnungs-ID oder Duftreferenz fehlt.")
            continue
        key, original, alternative = str(key), str(original), str(alternative)
        if key in twins:
            errors.append(f"Duftzwillinge!{row_number}: doppelte Zuordnungs-ID {key}.")
        if original == alternative:
            errors.append(f"Duftzwillinge!{row_number}: Original und Alternative sind identisch.")
        if original not in fragrances:
            errors.append(f"Duftzwillinge!{row_number}: Original-ID {original} unbekannt.")
        if alternative not in fragrances:
            errors.append(f"Duftzwillinge!{row_number}: Alternative-ID {alternative} unbekannt.")
        pair = (original, alternative)
        if pair in seen_pairs:
            errors.append(f"Duftzwillinge!{row_number}: doppelte Zuordnung {original} → {alternative}.")
        seen_pairs.add(pair)
        similarity = _float(data.get("DZI Ähnlichkeit (0–100)"))
        if similarity is None or not 0 <= similarity <= 100:
            errors.append(f"Duftzwillinge!{row_number}: ungültige Ähnlichkeit.")
        twins[key] = data

    if sheets["perfumers"] is not None:
        for row_number, data in _rows(sheets["perfumers"]):
            key, name = data.get("Parfümeur-ID"), data.get("Name")
            if not key or not name:
                errors.append(f"Parfümeure!{row_number}: Parfümeur-ID oder Name fehlt.")
                continue
            key = str(key)
            if key in perfumers:
                errors.append(f"Parfümeure!{row_number}: doppelte Parfümeur-ID {key}.")
            perfumers[key] = data

    if sheets["sources"] is not None:
        for row_number, data in _rows(sheets["sources"]):
            key, name = data.get("Quellen-ID"), data.get("Quellenname")
            if not key and not name:
                continue
            if not key or not name:
                errors.append(f"Quellen!{row_number}: Quellen-ID oder Quellenname fehlt.")
                continue
            key = str(key)
            if key in sources:
                errors.append(f"Quellen!{row_number}: doppelte Quellen-ID {key}.")
            sources[key] = data

    return brands, fragrances, twins, perfumers, sources, errors


def _build_plan(db: Session, workbook, sheets, filename: str, *, apply: bool) -> MasterImportReport:
    brand_rows, fragrance_rows, twin_rows, perfumer_rows, source_rows, errors = _validate(workbook, sheets)
    report = MasterImportReport(
        filename=filename,
        detected_version=_detect_version(workbook),
        dry_run=not apply,
        errors=errors,
    )
    if errors:
        raise MasterImportValidationError("\n".join(errors[:30]))
    if report.detected_version and not report.detected_version.startswith("2"):
        report.warnings.append(f"Erkannte Dateiversion {report.detected_version}; vorgesehen ist Master Database v2.x.")

    existing_brands = list(db.scalars(select(Brand)))
    brands_by_dgd = {item.dgd_id: item for item in existing_brands if item.dgd_id}
    brands_by_name = {_normal(item.name): item for item in existing_brands}
    resolved_brands: dict[str, Brand] = {}

    for master_id, data in brand_rows.items():
        existing = brands_by_dgd.get(master_id) or brands_by_name.get(_normal(data.get("Marke")))
        master_data = _json_row(data, workbook.epoch)
        if existing is None:
            report.brands.created += 1
            existing = Brand(
                id=uuid4(),
                dgd_id=master_id,
                name=str(data["Marke"]),
                country=data.get("Land"),
                founded_year=_int(data.get("Gründungsjahr") or data.get("Gruendungsjahr")),
                website_url=data.get("Website") or data.get("Offizielle Website"),
                verification_status=str(data.get("Verifizierungsstatus") or "OPEN").strip().upper(),
                description=data.get("Notizen"),
                master_data=master_data,
                active=str(data.get("Status") or "Aktiv").strip().casefold()
                not in {"inaktiv", "inactive", "nein", "no", "0", "false"},
            )
            if apply:
                db.add(existing)
                db.flush()
            resolved_brands[master_id] = existing
            brands_by_dgd[master_id] = existing
            brands_by_name[_normal(existing.name)] = existing
            continue

        changes: dict[str, Any] = {"dgd_id": master_id, "master_data": master_data}
        if data.get("Land") is not None:
            changes["country"] = data.get("Land")
        _fill_only(existing, "founded_year", _int(data.get("Gründungsjahr") or data.get("Gruendungsjahr")), changes)
        _fill_only(existing, "website_url", data.get("Website") or data.get("Offizielle Website"), changes)
        if data.get("Verifizierungsstatus") is not None:
            changes["verification_status"] = str(data.get("Verifizierungsstatus")).strip().upper()
        _fill_only(existing, "description", data.get("Notizen"), changes)
        if _changed(existing, changes):
            report.brands.updated += 1
            if apply:
                for key, value in changes.items():
                    setattr(existing, key, value)
        else:
            report.brands.unchanged += 1
        resolved_brands[master_id] = existing

    if apply:
        db.flush()

    existing_perfumers = {item.id: item for item in db.scalars(select(MasterPerfumer))}
    resolved_perfumers: dict[str, MasterPerfumer] = {}
    for master_id, data in perfumer_rows.items():
        existing = existing_perfumers.get(master_id)
        values = {
            "name": str(data["Name"]),
            "birth_year": _int(data.get("Geburtsjahr")),
            "nationality": data.get("Nationalität"),
            "profile": data.get("Karriere/Profil"),
            "style": data.get("Duftstil"),
            "notable_works": data.get("Bedeutende Werke"),
            "article_status": data.get("Artikelstatus"),
            "primary_source": data.get("Primärquelle"),
            "note": data.get("Notizen"),
            "master_data": _json_row(data, workbook.epoch),
        }
        if existing is None:
            report.perfumers.created += 1
            existing = MasterPerfumer(id=master_id, **values)
            if apply:
                db.add(existing)
            existing_perfumers[master_id] = existing
        elif _changed(existing, values):
            report.perfumers.updated += 1
            if apply:
                for key, value in values.items():
                    setattr(existing, key, value)
        else:
            report.perfumers.unchanged += 1
        resolved_perfumers[master_id] = existing

    if apply:
        db.flush()

    existing_fragrances = list(db.scalars(select(Fragrance)))
    fragrances_by_dgd = {item.dgd_id: item for item in existing_fragrances if item.dgd_id}
    exact_key: dict[tuple[Any, str, str], Fragrance] = {}
    name_key: dict[tuple[Any, str], list[Fragrance]] = {}
    for item in existing_fragrances:
        exact_key[(item.brand_id, _normal(item.name), _normal(item.concentration))] = item
        name_key.setdefault((item.brand_id, _normal(item.name)), []).append(item)
    resolved_fragrances: dict[str, Fragrance] = {}

    for master_id, data in fragrance_rows.items():
        brand = resolved_brands[str(data["Marken-ID"])]
        concentration = data.get("Konzentration")
        existing = fragrances_by_dgd.get(master_id)
        if existing is None:
            existing = exact_key.get((brand.id, _normal(data.get("Duftname")), _normal(concentration)))
        if existing is None:
            candidates = name_key.get((brand.id, _normal(data.get("Duftname"))), [])
            if len(candidates) == 1:
                existing = candidates[0]

        master_data = _json_row(data, workbook.epoch)
        perfumer_name = None
        perfumer_id = str(data.get("Parfümeur-ID")) if data.get("Parfümeur-ID") else None
        if perfumer_id and perfumer_id in resolved_perfumers:
            perfumer_name = resolved_perfumers[perfumer_id].name

        core_values = {
            "dgd_id": master_id,
            "name": str(data["Duftname"]),
            "brand_id": brand.id,
            "year": _int(data.get("Erscheinungsjahr")),
            "gender": _gender(data.get("Zielgruppe")),
            "concentration": concentration,
            "perfumer": perfumer_name,
            "image_url": data.get("Bild-URL") or data.get("Bild URL"),
            "image_source_name": data.get("Bildquelle") or data.get("Bildquelle Name"),
            "image_source_url": data.get("Bildquelle URL"),
            "image_usage_note": data.get("Bildrechte") or data.get("Bild Nutzungshinweis"),
            "image_status": str(data.get("Bildstatus") or "OPEN").strip().upper(),
            "master_data": master_data,
        }
        if existing is None:
            report.fragrances.created += 1
            item = Fragrance(
                id=uuid4(),
                **core_values,
                description=data.get("Kurznotiz"),
                accords=data.get("Duftfamilie"),
                projection=_float(data.get("Sillage (1–10)")),
            )
            if apply:
                db.add(item)
                db.flush()
            resolved_fragrances[master_id] = item
            fragrances_by_dgd[master_id] = item
            exact_key[(item.brand_id, _normal(item.name), _normal(item.concentration))] = item
            name_key.setdefault((item.brand_id, _normal(item.name)), []).append(item)
            continue

        changes = dict(core_values)
        _fill_only(existing, "description", data.get("Kurznotiz"), changes)
        _fill_only(existing, "accords", data.get("Duftfamilie"), changes)
        _fill_only(existing, "projection", _float(data.get("Sillage (1–10)")), changes)
        if _changed(existing, changes):
            report.fragrances.updated += 1
            if apply:
                for key, value in changes.items():
                    setattr(existing, key, value)
        else:
            report.fragrances.unchanged += 1
        resolved_fragrances[master_id] = existing

    if apply:
        db.flush()

    existing_twins = list(db.scalars(select(TwinMatch)))
    twins_by_dgd = {item.dgd_id: item for item in existing_twins if item.dgd_id}
    twins_by_pair = {(item.original_id, item.alternative_id): item for item in existing_twins}

    for master_id, data in twin_rows.items():
        original = resolved_fragrances[str(data["Original-DGD-ID"])]
        alternative = resolved_fragrances[str(data["Alternative-DGD-ID"])]
        existing = twins_by_dgd.get(master_id) or twins_by_pair.get((original.id, alternative.id))
        profile = data.get("Duftprofil") or data.get("Kurzfazit")
        differences_parts = []
        if data.get("Opening"):
            differences_parts.append(f"Opening: {data['Opening']}")
        if data.get("Drydown"):
            differences_parts.append(f"Drydown: {data['Drydown']}")
        if data.get("Performance vs. Original"):
            differences_parts.append(f"Performance: {data['Performance vs. Original']}")
        source_parts = [
            value for value in (
                data.get("Quelle 1"), data.get("Quelle 2"), data.get("Vertrauensstatus"), data.get("Prüfstatus")
            ) if value
        ]
        master_data = _json_row(data, workbook.epoch)
        core_values = {
            "dgd_id": master_id,
            "original_id": original.id,
            "alternative_id": alternative.id,
            "similarity": _float(data.get("DZI Ähnlichkeit (0–100)")) or 0,
            "master_data": master_data,
        }
        if existing is None:
            report.twins.created += 1
            item = TwinMatch(
                id=uuid4(),
                **core_values,
                commonalities=profile,
                differences=" · ".join(differences_parts) or None,
                source_note=" · ".join(str(part) for part in source_parts) or None,
            )
            if apply:
                db.add(item)
            twins_by_dgd[master_id] = item
            twins_by_pair[(original.id, alternative.id)] = item
            continue

        changes = dict(core_values)
        _fill_only(existing, "commonalities", profile, changes)
        _fill_only(existing, "differences", " · ".join(differences_parts) or None, changes)
        _fill_only(existing, "source_note", " · ".join(str(part) for part in source_parts) or None, changes)
        if _changed(existing, changes):
            report.twins.updated += 1
            if apply:
                for key, value in changes.items():
                    setattr(existing, key, value)
        else:
            report.twins.unchanged += 1

    existing_sources = {item.id: item for item in db.scalars(select(MasterSource))}
    for master_id, data in source_rows.items():
        existing = existing_sources.get(master_id)
        values = {
            "name": str(data["Quellenname"]),
            "object_type": data.get("Objekttyp"),
            "object_id": data.get("Objekt-ID"),
            "source_type": data.get("Quellentyp"),
            "file_or_url": data.get("Datei/URL"),
            "source_date": _datetime(data.get("Stand"), workbook.epoch),
            "usage_status": data.get("Nutzungsstatus"),
            "trust_status": data.get("Vertrauensstatus"),
            "note": data.get("Hinweis"),
            "master_data": _json_row(data, workbook.epoch),
        }
        if existing is None:
            report.sources.created += 1
            if apply:
                db.add(MasterSource(id=master_id, **values))
        elif _changed(existing, values):
            report.sources.updated += 1
            if apply:
                for key, value in values.items():
                    setattr(existing, key, value)
        else:
            report.sources.unchanged += 1

    return report


def preview_master_import(db: Session, file_bytes: bytes, filename: str) -> dict[str, Any]:
    workbook, sheets = _prepare_workbook(file_bytes, filename)
    report = _build_plan(db, workbook, sheets, filename, apply=False)
    return report.to_dict()


def commit_master_import(db: Session, file_bytes: bytes, filename: str) -> dict[str, Any]:
    workbook, sheets = _prepare_workbook(file_bytes, filename)
    try:
        report = _build_plan(db, workbook, sheets, filename, apply=True)
        payload = report.to_dict()
        db.add(MasterImportRun(
            filename=filename,
            file_version=report.detected_version,
            status="completed",
            report=payload,
        ))
        db.commit()
        return payload
    except Exception:
        db.rollback()
        raise
